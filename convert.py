import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import csv


def main(args):
    path = args[0]
    subdirs = [x[0] for x in os.walk(path)]
    for subdir_1 in subdirs[1:]:
        print(subdir_1)
        subdirs_2 = [x[0] for x in os.walk(subdir_1)]
        for subdir in subdirs_2[1:]:
            print(subdir)
            arr = [f.split(".") for f in os.listdir(subdir)]
            files = []
            for name in arr:
                print(name)
                if name[1] == "csv":
                    files.append(name[0])
            for filename in files:
                file_name  = os.path.join(path,subdir,filename + '.csv')
                f = open(file_name, 'rt')
                csv.register_dialect('commas', delimiter=',')
                reader = csv.reader(f, dialect='commas')
                wb = Workbook()
                new_file_name = os.path.join(path,subdir,filename + '.xlsx')
                dest_filename = new_file_name
                ws = wb.worksheets[0]
                i = 0
                for row_index, row in enumerate(reader):
                    if row_index < 4:
                        for column_index, cell in enumerate(row):
                            column_letter = get_column_letter((column_index + 1))
                            ws.cell(column=(column_index + 1),row=(row_index + 1),value=cell)
                    else:
                        for column_index, cell in enumerate(row):
                            column_letter = get_column_letter((column_index + 1))
                            ws.cell(column=(column_index + 1),row=(row_index + 1),value=float(cell))
                sheet = wb.worksheets[0]
                sheet.delete_rows(1,4)
                wb.save(new_file_name)


if __name__ == "__main__":
    main(sys.argv[1:])
